import streamlit as st
import os
import csv
import time
import datetime
import numpy as np


import torch
import torch.nn as nn
from torch.nn.utils import prune
import torch.nn.functional as F
import torch.optim as optim
from torch.utils.data import  DataLoader
from torchvision import transforms
from torch.amp import GradScaler, autocast
from torch.utils.tensorboard import SummaryWriter
from openpyxl import Workbook, load_workbook

# --- Import Model Architectures ---
from utility.models.vanila_model import ConvNeXt6DP, PoseDataset
from utility.models.six_ch_model import ConvNeXt6DP6ch, Stereo6ChPoseDataset
from utility.models.sw_twin_model import StereoConvNeXt6DP, StereoPoseDataset

# --- Import Loss and Metrics functions ---
from utility.loss_metrics_utils import *

# --- Configuration and Paths ---
# Base directory for datasets, models, and logs.
# This now points to the parent directory of 'streamlit'
BASE_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))

# Ensure necessary directories exist relative to BASE_DIR
os.makedirs(os.path.join(BASE_DIR, "model"), exist_ok=True)
os.makedirs(os.path.join(BASE_DIR, "runs"), exist_ok=True)
os.makedirs(os.path.join(BASE_DIR, "dataset"), exist_ok=True) # Ensure dataset base exists

DEVICE = torch.device("cuda" if torch.cuda.is_available() else "cpu")

# --- DataLoader Setup ---
@st.cache_resource
def get_transform(img_size):
    """Returns the image transformation pipeline."""
    return transforms.Compose([
        transforms.Resize((img_size, img_size)),
        transforms.ToTensor(),
        transforms.Normalize(
            mean=[0.485, 0.456, 0.406],  # ImageNet mean
            std=[0.229, 0.224, 0.225]    # ImageNet std
        )
    ])

@st.cache_resource
def get_dataloader(model_type, batch_id, img_size, batch_size, split):
    """
    Returns a DataLoader for the specified model type and split.
    Uses st.cache_resource to avoid reloading data unnecessarily.
    """
    DATASET_DIR = os.path.join(BASE_DIR, f"dataset/batch{batch_id}")
    base_dir = os.path.join(DATASET_DIR, split)
    csv_path = os.path.join(base_dir, "labels.csv")
    images_dir = os.path.join(base_dir, "images")

    if not os.path.exists(csv_path) or not os.path.exists(images_dir):
        st.error(f"Dataset for batch {batch_id} at '{DATASET_DIR}' not found. Please ensure the structure is correct: {DATASET_DIR}/train/images, {DATASET_DIR}/train/labels.csv, etc.")
        st.stop() # Stop execution if dataset is missing

    transform = get_transform(img_size)
    dataset = None
    if model_type == "Vanilla":
        dataset = PoseDataset(csv_path, images_dir, transform=transform)
    elif model_type == "Stereo Shared Weights (Twin Heads)":
        dataset = StereoPoseDataset(csv_path, images_dir, transform=transform)
    elif model_type == "6-Channel Stacked":
        dataset = Stereo6ChPoseDataset(csv_path, images_dir, transform=transform)
    else:
        st.error(f"Unknown model type: {model_type}")
        st.stop()

    return DataLoader(dataset, batch_size=batch_size, shuffle=(split == "train"))

# --- Model Instantiation Function ---
def get_model_architecture(model_name, head_layers):
    """
    Returns an instance of the selected model architecture.
    """
    if model_name == "Vanilla":
        return ConvNeXt6DP(head_layers=head_layers).to(DEVICE)
    elif model_name == "Stereo Shared Weights (Twin Heads)":
        return StereoConvNeXt6DP(head_layers=head_layers).to(DEVICE)
    elif model_name == "6-Channel Stacked":
        return ConvNeXt6DP6ch(head_layers=head_layers, in_chans=6).to(DEVICE)
    else:
        raise ValueError(f"Unknown model architecture: {model_name}")

# --- Training Function ---
def train(model, optimizer, scheduler, train_loader, val_loader,
          num_epochs, patience, trans_weight, rotation_weight,
          model_save_path, best_model_path, progress_bar, status_text):
    """
    Trains the model and handles validation, early stopping, and checkpointing.
    """
    writer = SummaryWriter(log_dir=os.path.join(BASE_DIR, f"runs/{model.__class__.__name__}_{datetime.datetime.now().strftime('%Y%m%d-%H%M%S')}"))
    scaler = GradScaler()
    best_val_loss = float('inf')
    epochs_no_improve = 0
    start_epoch = 0
    time_per_epoch = []

    status_text.info("Starting training...")

    for epoch in range(start_epoch, num_epochs):
        epoch_start_time = time.time()
        status_text.info(f"EPOCH : {epoch + 1}/{num_epochs}")
        model.train()
        train_loss, total_trans_rmse, total_rot_rmse = 0.0, 0.0, 0.0
        num_samples = 0

        # Use Streamlit progress bar for the training loop
        batch_progress = st.progress(0, text="Batch Progress")
        for i, batch_data in enumerate(train_loader):
            if model.__class__.__name__ == "StereoConvNeXt6DP":
                imagesL, imagesR, labels = batch_data
                imagesL = imagesL.to(DEVICE)
                imagesR = imagesR.to(DEVICE)
            else:
                images, labels = batch_data
                images = images.to(DEVICE)
            labels = labels.to(DEVICE)
            batch_size = labels.size(0)

            optimizer.zero_grad(set_to_none=True)
            with autocast(device_type="cuda"):
                if model.__class__.__name__ == "StereoConvNeXt6DP":
                    outputs = model(imagesL, imagesR)
                else:
                    outputs = model(images)
                loss = combined_loss(outputs, labels, trans_weight, rotation_weight)

            scaler.scale(loss).backward()
            scaler.step(optimizer)
            scaler.update()
            torch.cuda.synchronize()

            train_loss += loss.item() * batch_size
            trans_rmse, rot_rmse = compute_errors(outputs, labels)
            total_trans_rmse += trans_rmse * batch_size
            total_rot_rmse += rot_rmse * batch_size
            num_samples += batch_size

            batch_progress.progress((i + 1) / len(train_loader), text=f"Batch Progress: Loss {loss.item():.4f}")

        avg_train_loss = train_loss / num_samples
        avg_trans_rmse = total_trans_rmse / num_samples
        avg_rot_rmse = total_rot_rmse / num_samples

        epoch_end_time = time.time()
        epoch_duration = int(epoch_end_time - epoch_start_time)
        time_per_epoch.append(epoch_duration)

        st.write(f"✅ Avg Training Loss: {avg_train_loss:.4f}")
        st.write(f"📊 RMSE - Trans: {avg_trans_rmse:.4f}, Rot: {avg_rot_rmse:.4f}")
        st.write(f"⏱️ Time per epoch: {epoch_duration}s")

        writer.add_scalar("Loss/Train", avg_train_loss, epoch)
        writer.add_scalar("RMSE/Train_Translation", avg_trans_rmse, epoch)
        writer.add_scalar("RMSE/Train_Rotation", avg_rot_rmse, epoch)

        # Pruning logic (every 5 epochs, excluding epoch 0)
        if epoch != 0 and (epoch + 1) % 5 == 0:
            parameters_to_prune = [
                (module, 'weight') for module in model.modules()
                if isinstance(module, (nn.Linear, nn.Conv2d)) and hasattr(module, 'weight')
            ]
            if parameters_to_prune:
                prune.global_unstructured(
                    parameters_to_prune,
                    pruning_method=prune.L1Unstructured,
                    amount=0.1
                )
                st.warning(f"⚠️ Pruning applied at epoch {epoch + 1}")
            else:
                st.warning(f"⚠️ Skipping pruning: No eligible parameters found at epoch {epoch + 1}")

        # Validation
        model.eval()
        val_loss, val_trans_rmse, val_rot_rmse = 0.0, 0.0, 0.0
        val_samples = 0
        with torch.no_grad():
            for batch_data in val_loader:
                if model.__class__.__name__ == "StereoConvNeXt6DP":
                    imagesL, imagesR, labels = batch_data
                    imagesL = imagesL.to(DEVICE)
                    imagesR = imagesR.to(DEVICE)
                else:
                    images, labels = batch_data
                    images = images.to(DEVICE)
                labels = labels.to(DEVICE)
                batch_size = labels.size(0)

                with autocast(device_type="cuda"):
                    if model.__class__.__name__ == "StereoConvNeXt6DP":
                        outputs = model(imagesL, imagesR)
                    else:
                        outputs = model(images)
                    loss = combined_loss(outputs, labels, trans_weight, rotation_weight)
                torch.cuda.synchronize()
                val_loss += loss.item() * batch_size
                trans_rmse, rot_rmse = compute_errors(outputs, labels)
                val_trans_rmse += trans_rmse * batch_size
                val_rot_rmse += rot_rmse * batch_size
                val_samples += batch_size

        avg_val_loss = val_loss / val_samples
        avg_val_trans_rmse = val_trans_rmse / val_samples
        avg_val_rot_rmse = val_rot_rmse / val_samples

        st.write(f"⭐ Val Loss: {avg_val_loss:.4f} | RMSE Trans: {avg_val_trans_rmse:.4f}, Rot: {avg_val_rot_rmse:.4f}")

        writer.add_scalar("Loss/Val", avg_val_loss, epoch)
        writer.add_scalar("RMSE/Val_Translation", avg_val_trans_rmse, epoch)
        writer.add_scalar("RMSE/Val_Rotation", avg_val_rot_rmse, epoch)

        scheduler.step(avg_val_loss)

        if avg_val_loss < best_val_loss:
            best_val_loss = avg_val_loss
            epochs_no_improve = 0
            torch.save({
                'model_state_dict': model.state_dict(),
                'optimizer_state_dict': optimizer.state_dict(),
                'scaler_state_dict': scaler.state_dict(),
                'best_val_loss': best_val_loss,
                'epochs_no_improve': epochs_no_improve,
                'epoch': epoch + 1
            }, best_model_path)
            st.success(f"🏆 Best model saved to: {best_model_path}")
        else:
            epochs_no_improve += 1
            st.warning(f"⚠️ No improvement ({epochs_no_improve}/{patience})")
            if epochs_no_improve >= patience:
                st.error("⏹️ Early stopping triggered")
                break

        # Save checkpoint after each epoch
        torch.save({
            'model_state_dict': model.state_dict(),
            'optimizer_state_dict': optimizer.state_dict(),
            'scaler_state_dict': scaler.state_dict(),
            'best_val_loss': best_val_loss,
            'epochs_no_improve': epochs_no_improve,
            'epoch': epoch + 1
        }, model_save_path)
        st.info(f"💾 Checkpoint saved to: {model_save_path}")

    writer.close()
    return time_per_epoch

# --- Evaluation Functions (Test and Validate) ---
def evaluate_model(model, loader, model_type, mode='Test', use_amp=True):
    """
    Evaluates the model on a given DataLoader.
    """
    model.eval()
    inference_times = []
    total_loss = 0.0
    total_trans_rmse = 0.0
    total_rot_rmse = 0.0
    num_samples = 0
    all_preds, all_gts = [], []

    eval_progress_bar = st.progress(0, text=f"Running {mode}...")
    with torch.no_grad():
        for i, batch_data in enumerate(loader):
            if model_type == "Stereo Shared Weights (Twin Heads)":
                imagesL, imagesR, labels = batch_data
                imagesL = imagesL.to(DEVICE)
                imagesR = imagesR.to(DEVICE)
            else:
                images, labels = batch_data
                images = images.to(DEVICE)
            labels = labels.to(DEVICE)
            batch_size = labels.size(0)

            torch.cuda.synchronize()
            start_time = time.time()

            if use_amp:
                with autocast(device_type="cuda"):
                    if model_type == "Stereo Shared Weights (Twin Heads)":
                        outputs = model(imagesL, imagesR)
                    else:
                        outputs = model(images)
                    loss = combined_loss(outputs, labels)
            else:
                if model_type == "Stereo Shared Weights (Twin Heads)":
                    outputs = model(imagesL, imagesR)
                else:
                    outputs = model(images)
                loss = combined_loss(outputs, labels)

            torch.cuda.synchronize()
            inference_time = (time.time() - start_time) * 1000 / batch_size  # ms per image
            inference_times.append(inference_time)

            total_loss += loss.item() * batch_size
            trans_rmse, rot_rmse = compute_errors(outputs, labels)
            total_trans_rmse += trans_rmse * batch_size
            total_rot_rmse += rot_rmse * batch_size
            num_samples += batch_size

            all_preds.append(outputs.cpu().numpy())
            all_gts.append(labels.cpu().numpy())
            eval_progress_bar.progress((i + 1) / len(loader), text=f"Running {mode}...")

    avg_loss = total_loss / num_samples
    avg_trans_rmse = total_trans_rmse / num_samples
    avg_rot_rmse = total_rot_rmse / num_samples
    avg_inference_time = np.mean(inference_times)

    return avg_loss, avg_trans_rmse, avg_rot_rmse, np.concatenate(all_preds), np.concatenate(all_gts), avg_inference_time

# --- Logging Functions ---
def write_md_report(batch_id, mod_id, config, metrics, dataset_stats, file_paths):
    """Writes an evaluation report to a Markdown file."""
    eval_path = file_paths['eval_md_path']
    eval_content = f"""# Evaluation Results - Batch {batch_id} - Model {mod_id}

## Training Configuration
- Model Architecture: {config['model_architecture']}
- Batch Size: {config['batch_size']}
- Epochs: {config['num_epochs']}
- Learning Rate: {config['learning_rate']}
- Translation Weight : {config['trans_weight']}
- Rotation Weight : {config['rotation_weight']}
- Angular Weight : {config['angular_weight']}
- Patience : {config['patience']}
- Image Size: {config['img_size']}
- Device: {DEVICE}
- Optimizer : Adam
- Head Layers: {config['head_layers']}

## Evaluation Metrics

### Test Set
- Average Loss: {metrics['test_avg_loss']:.4f}
- Translation RMSE: {metrics['test_avg_trans_rmse']:.4f}
- Translation Accuracy: {metrics['test_translation_rmse_cm']:.2f} cm
- Translation Accuracy %: {metrics['test_trans_accuracy_pct']:.2f}%
- Rotation RMSE: {metrics['test_avg_rot_rmse']:.4f}
- Rotation Accuracy: {metrics['test_rot_accuracy']:.2f}°
- Rotation Accuracy % : {metrics['test_rot_accuracy_pct']:.2f} %
- Inference Speed: {metrics['test_avg_inference_time']:.2f} ms/frame

### Validation Set
- Average Loss: {metrics['val_avg_loss']:.4f}
- Translation RMSE: {metrics['val_avg_trans_rmse']:.4f}
- Translation Accuracy: {metrics['val_translation_rmse_cm']:.2f} cm
- Translation Accuracy %: {metrics['val_trans_accuracy_pct']:.2f}%
- Rotation RMSE: {metrics['val_avg_rot_rmse']:.4f}
- Rotation Accuracy: {metrics['val_rot_accuracy']:.2f}°
- Rotation Accuracy % : {metrics['val_rot_accuracy_pct']:.2f} %
- Inference Speed: {metrics['val_avg_inference_time']:.2f} ms/frame

## Dataset Statistics
### Training Set
- Translation range: [{dataset_stats['train_min_mean']:.2f}, {dataset_stats['train_max_mean']:.2f}] m

### Validation Set
- Translation range: [{dataset_stats['val_min_mean']:.2f}, {dataset_stats['val_max_mean']:.2f}] m

### Test Set
- Translation range: [{dataset_stats['test_min_mean']:.2f}, {dataset_stats['test_max_mean']:.2f}] m

## File Locations
- Dataset Directory: {file_paths['dataset_dir']}
- Model Save Path: {file_paths['model_save_path']}
- Best Model Path: {file_paths['best_model_path']}
- Clean Model Path: {file_paths['clean_model_path']}
"""
    with open(eval_path, 'w') as f:
        f.write(eval_content)
    st.success(f"📄 Evaluation report saved to: {eval_path}")

def append_to_csv(batch_id, mod_id, config, metrics, file_paths):
    """Appends evaluation results to a CSV file."""
    csv_path = os.path.join(BASE_DIR, "model/eval_results.csv")
    write_header = not os.path.exists(csv_path)

    csv_data = {
        'timestamp': datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'dataset_id': batch_id,
        'model_id': mod_id,
        'model_architecture': config['model_architecture'],
        'head_layers': str(config['head_layers']),
        'batch_size': config['batch_size'],
        'epochs': config['num_epochs'],
        'learning_rate': config['learning_rate'],
        'translation_weight': config['trans_weight'],
        'rotation_weight': config['rotation_weight'],
        'angular_weight': config['angular_weight'],
        'patience': config['patience'],
        'img_size': config['img_size'],
        'test_loss': metrics['test_avg_loss'],
        'test_translation_rmse': metrics['test_avg_trans_rmse'],
        'test_translation_accuracy_pct': metrics['test_trans_accuracy_pct'],
        'test_rotation_rmse': metrics['test_avg_rot_rmse'],
        'test_rotation_accuracy_pct': metrics['test_rot_accuracy_pct'],
        'test_inference_time_ms': metrics['test_avg_inference_time'],
        'validation_loss': metrics['val_avg_loss'],
        'validation_translation_rmse': metrics['val_avg_trans_rmse'],
        'validation_translation_accuracy_pct': metrics['val_trans_accuracy_pct'],
        'validation_rotation_rmse': metrics['val_avg_rot_rmse'],
        'validation_rotation_accuracy_pct': metrics['val_rot_accuracy_pct'],
        'validation_inference_time_ms': metrics['val_avg_inference_time'],
        'model_path': file_paths['model_save_path'],
        'eval_path': file_paths['eval_md_path']
    }

    with open(csv_path, 'a', newline='') as csvfile:
        writer = csv.DictWriter(csvfile, fieldnames=csv_data.keys())
        if write_header:
            writer.writeheader()
        writer.writerow(csv_data)
    st.success(f"📊 Results appended to CSV: {csv_path}")

def log_eval_to_xlsx(batch_id, mod_id, config, metrics, dataset_stats, file_paths, train_time_minutes, param_count, notes):
    """Appends evaluation results to an Excel file."""
    xlsx_path = os.path.join(BASE_DIR, "model/EVAL.xlsx")
    headers = [
        "ID", "IMG_SIZE", "VARIANT", "HEAD_ARCH", "PARAM_COUNT", "BATCH_SIZE", "EPC", "DTS_ID", "DTS_LEN", "LR_RT",
        "TRAIN_TIME", "VRAM_USG", "VAL_LOSS", "TS_LOSS",
        "VAL_TRANS_RSME", "TS_TRANS_RMSE", "VAL_ROT_RSME", "TS_ROT_RMSE",
        "VAL_TRANS_ACC", "TS_TRANS_ACC", "VAL_ROT_ACC", "TS_ROT_ACC",
        "VAL_INF_MS", "TS_INF_MS", "Notes"
    ]

    model_id_str = f"{batch_id}{mod_id}"
    dataset_len = (len(st.session_state.train_loader.dataset) +
                   len(st.session_state.val_loader.dataset) +
                   len(st.session_state.test_loader.dataset))

    # Placeholder for VRAM usage - actual measurement is complex in Streamlit/iFrame
    vram_usage_gb = 0.0 # Will try to get actual if possible, otherwise keep as placeholder

    row = [
        int(model_id_str), config['img_size'], config['model_architecture'], str(config['head_layers']), param_count, config['batch_size'], config['num_epochs'],
        batch_id, dataset_len, str(config['learning_rate']),
        round(train_time_minutes, 1), vram_usage_gb, # TRAIN_TIME, VRAM_USG

        round(metrics['val_avg_loss'], 4), round(metrics['test_avg_loss'], 4),
        round(metrics['val_avg_trans_rmse'], 4),
        round(metrics['test_avg_trans_rmse'], 4),
        round(metrics['val_avg_rot_rmse'], 4),
        round(metrics['test_avg_rot_rmse'], 4),

        round(metrics['val_trans_accuracy_pct'], 2), round(metrics['test_trans_accuracy_pct'], 2),
        round(metrics['val_rot_accuracy_pct'], 2), round(metrics['test_rot_accuracy_pct'], 2),

        round(metrics['val_avg_inference_time'], 2), round(metrics['test_avg_inference_time'], 2),

        notes
    ]

    row = [safe(x) for x in row]

    # Load or create workbook
    if os.path.exists(xlsx_path):
        wb = load_workbook(xlsx_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "ConvNeXt V2 Nano Evaluation"
        ws.append(headers)

    ws.append(row)
    wb.save(xlsx_path)
    st.success(f"✅ Evaluation logged to Excel: {xlsx_path}")


# --- Streamlit UI ---
st.set_page_config(layout="wide", page_title="6D Pose Estimation Model Trainer")

st.title("6D Pose Estimation Model Trainer")

st.markdown("""
This application allows you to train, validate, and test different 6D pose estimation model architectures
with customizable parameters. Results will be saved to Markdown, CSV, and Excel files.
""")

# --- Input Parameters ---
st.header("Configuration")

col1, col2, col3 = st.columns(3)

with col1:
    model_architecture = st.selectbox(
        "Select Model Architecture",
        ("Vanilla", "Stereo Shared Weights (Twin Heads)", "6-Channel Stacked"),
        help="Choose the model variant to train."
    )
    batch_id = st.number_input("Dataset Batch ID", min_value=1, value=5, step=1, help="Identifier for the dataset batch (e.g., 'batch5').")
    mod_id = st.number_input("Model ID", min_value=1, value=2, step=1, help="Unique identifier for this specific model run.")
    img_size = st.number_input("Image Size (pixels)", min_value=64, value=224, step=32, help="Input image resolution (e.g., 224x224).")

with col2:
    num_epochs = st.number_input("Number of Epochs", min_value=1, value=30, step=5, help="Total training epochs.")
    batch_size = st.number_input("Batch Size", min_value=1, value=32, step=4, help="Number of samples per batch.")
    learning_rate = st.number_input("Learning Rate", min_value=1e-6, max_value=1e-2, value=1e-4, format="%.6f", help="Optimizer learning rate.")
    patience = st.number_input("Early Stopping Patience", min_value=1, value=3, step=1, help="Number of epochs with no improvement after which training will be stopped.")

with col3:
    trans_weight = st.number_input("Translation Loss Weight", min_value=0.1, value=1.5, step=0.1, help="Weight for translation component in combined loss.")
    rotation_weight = st.number_input("Rotation Loss Weight", min_value=0.1, value=1.0, step=0.1, help="Weight for rotation component in combined loss.")
    angular_weight = st.number_input("Angular Loss Weight (Not Used in Combined Loss)", min_value=0.0, value=0.1, step=0.1, help="This parameter is present in your original code but not used in the combined_loss function. Keeping for consistency.")
    head_layers_str = st.text_input("Head Layers (comma-separated integers)", value="512", help="Define the hidden layer sizes for the model's head (e.g., '512,256').")
    notes = st.text_area("Notes for Evaluation Log", "Initial run with selected parameters.", height=100)

# --- Start Training Button ---
st.markdown("---")
if st.button("🚀 Start Training"):
    st.header("Training Progress")
    progress_placeholder = st.empty()
    status_placeholder = st.empty()
    st.write("---") # Separator for metrics

    try:
        # Parse head layers
        head_layers = [int(x.strip()) for x in head_layers_str.split(',') if x.strip()]
        if not head_layers:
            st.error("Head Layers cannot be empty. Please enter at least one layer size.")
            st.stop()

        # Define file paths for this run
        model_name_prefix = model_architecture.replace(" ", "").replace("(", "").replace(")", "").replace("-", "")
        MODEL_SAVE_PATH = os.path.join(BASE_DIR, f"model/{model_name_prefix}{batch_id}.{mod_id}.pth")
        BEST_MODEL_PATH = os.path.join(BASE_DIR, f"model/BEST-{model_name_prefix}{batch_id}.{mod_id}.pth")
        CLEAN_MODEL_PATH = os.path.join(BASE_DIR, f"model/CLEAN-{model_name_prefix}{batch_id}.{mod_id}.pth")
        EVAL_MD_PATH = os.path.join(BASE_DIR, f"model/{model_name_prefix}{batch_id}.{mod_id}.md")

        file_paths = {
            'dataset_dir': os.path.join(BASE_DIR, f"dataset/batch{batch_id}"),
            'model_save_path': MODEL_SAVE_PATH,
            'best_model_path': BEST_MODEL_PATH,
            'clean_model_path': CLEAN_MODEL_PATH,
            'eval_md_path': EVAL_MD_PATH
        }

        # Store loaders in session state to avoid re-creating for stats later
        st.session_state.train_loader = get_dataloader(model_architecture, batch_id, img_size, batch_size, "train")
        st.session_state.val_loader = get_dataloader(model_architecture, batch_id, img_size, batch_size, "val")
        st.session_state.test_loader = get_dataloader(model_architecture, batch_id, img_size, batch_size, "test")

        # Initialize Model
        model = get_model_architecture(model_architecture, head_layers)
        optimizer = optim.Adam(model.parameters(), lr=learning_rate)
        scheduler = optim.lr_scheduler.ReduceLROnPlateau(optimizer, patience=patience)

        # Calculate parameter count
        param_count = sum(p.numel() for p in model.parameters() if p.requires_grad)
        st.info(f"Model {model_architecture} initialized with {param_count:,} trainable parameters.")

        # --- Train Model ---
        status_placeholder.info("Starting model training...")
        time_per_epoch = train(
            model, optimizer, scheduler,
            st.session_state.train_loader, st.session_state.val_loader,
            num_epochs, patience, trans_weight, rotation_weight,
            MODEL_SAVE_PATH, BEST_MODEL_PATH, progress_placeholder, status_placeholder
        )
        total_train_time_minutes = sum(time_per_epoch) / 60.0
        status_placeholder.success("Training complete!")

        # --- Save Clean Model ---
        status_placeholder.info("Saving clean model (removing pruning reparametrizations)...")
        for name, module in model.named_modules():
            if isinstance(module, torch.nn.Conv2d) or isinstance(module, torch.nn.Linear):
                if hasattr(module, "weight_orig"):
                    prune.remove(module, "weight")
        torch.save(model.state_dict(), CLEAN_MODEL_PATH)
        status_placeholder.success(f"Clean model saved to: {CLEAN_MODEL_PATH}")

        # --- Test Model ---
        status_placeholder.info("Running final test evaluation...")
        test_avg_loss, test_avg_trans_rmse, test_avg_rot_rmse, test_preds, test_gts, test_avg_inference_time = evaluate_model(
            model, st.session_state.test_loader, model_architecture, mode='Test', use_amp=True
        )
        status_placeholder.success("Test evaluation complete!")

        # --- Validate Model (on clean model) ---
        status_placeholder.info("Running final validation evaluation...")
        val_avg_loss, val_avg_trans_rmse, val_avg_rot_rmse, val_preds, val_gts, val_avg_inference_time = evaluate_model(
            model, st.session_state.val_loader, model_architecture, mode='Validation', use_amp=True
        )
        status_placeholder.success("Validation evaluation complete!")

        # --- Calculate Final Metrics ---
        status_placeholder.info("Calculating final metrics...")
        test_translation_rmse_cm = calculate_translation_rmse(test_preds, test_gts)
        val_translation_rmse_cm = calculate_translation_rmse(val_preds, val_gts)
        test_rot_accuracy = rotation_error_deg_from_6d(torch.tensor(test_preds[:, 3:]), torch.tensor(test_gts[:, 3:]))
        val_rot_accuracy = rotation_error_deg_from_6d(torch.tensor(val_preds[:, 3:]), torch.tensor(val_gts[:, 3:]))

        train_trans_stats = get_dataset_stats(st.session_state.train_loader)
        val_trans_stats = get_dataset_stats(st.session_state.val_loader)
        test_trans_stats = get_dataset_stats(st.session_state.test_loader)

        test_range_cm = (test_trans_stats['max'].mean() - test_trans_stats['min'].mean()) * 100
        val_range_cm = (val_trans_stats['max'].mean() - val_trans_stats['min'].mean()) * 100

        test_trans_accuracy_pct = translation_accuracy_percentage(test_translation_rmse_cm, test_range_cm).item()
        val_trans_accuracy_pct = translation_accuracy_percentage(val_translation_rmse_cm, val_range_cm).item()

        test_rot_accuracy_pct = 100 * (1 - (test_rot_accuracy.item() / 360))
        val_rot_accuracy_pct = 100 * (1 - (val_rot_accuracy.item() / 360))
        status_placeholder.success("Metrics calculated!")

        # Consolidate metrics and config for logging functions
        metrics_dict = {
            'test_avg_loss': test_avg_loss,
            'test_avg_trans_rmse': test_avg_trans_rmse,
            'test_translation_rmse_cm': test_translation_rmse_cm,
            'test_trans_accuracy_pct': test_trans_accuracy_pct,
            'test_avg_rot_rmse': test_avg_rot_rmse,
            'test_rot_accuracy': test_rot_accuracy.item(),
            'test_rot_accuracy_pct': test_rot_accuracy_pct,
            'test_avg_inference_time': test_avg_inference_time,
            'val_avg_loss': val_avg_loss,
            'val_avg_trans_rmse': val_avg_trans_rmse,
            'val_translation_rmse_cm': val_translation_rmse_cm,
            'val_trans_accuracy_pct': val_trans_accuracy_pct,
            'val_avg_rot_rmse': val_avg_rot_rmse,
            'val_rot_accuracy': val_rot_accuracy.item(),
            'val_rot_accuracy_pct': val_rot_accuracy_pct,
            'val_avg_inference_time': val_avg_inference_time,
        }

        config_dict = {
            'model_architecture': model_architecture,
            'batch_size': batch_size,
            'num_epochs': num_epochs,
            'learning_rate': learning_rate,
            'trans_weight': trans_weight,
            'rotation_weight': rotation_weight,
            'angular_weight': angular_weight,
            'patience': patience,
            'img_size': img_size,
            'head_layers': head_layers_str # Keep as string for MD/CSV/XLSX
        }

        dataset_stats_dict = {
            'train_min_mean': train_trans_stats['min'].mean().item(),
            'train_max_mean': train_trans_stats['max'].mean().item(),
            'val_min_mean': val_trans_stats['min'].mean().item(),
            'val_max_mean': val_trans_stats['max'].mean().item(),
            'test_min_mean': test_trans_stats['min'].mean().item(),
            'test_max_mean': test_trans_stats['max'].mean().item(),
        }

        # --- Write Reports ---
        status_placeholder.info("Writing evaluation reports...")
        write_md_report(batch_id, mod_id, config_dict, metrics_dict, dataset_stats_dict, file_paths)
        append_to_csv(batch_id, mod_id, config_dict, metrics_dict, file_paths)
        log_eval_to_xlsx(batch_id, mod_id, config_dict, metrics_dict, dataset_stats_dict, file_paths, total_train_time_minutes, param_count, notes)
        status_placeholder.success("All reports generated and saved!")

        st.balloons()
        st.header("Results Summary")
        col_res1, col_res2 = st.columns(2)
        with col_res1:
            st.metric("Test Translation RMSE (cm)", f"{test_translation_rmse_cm:.2f}")
            st.metric("Test Rotation Accuracy (°)", f"{test_rot_accuracy:.2f}")
            st.metric("Test Inference Time (ms/image)", f"{test_avg_inference_time:.2f}")
        with col_res2:
            st.metric("Validation Translation RMSE (cm)", f"{val_translation_rmse_cm:.2f}")
            st.metric("Validation Rotation Accuracy (°)", f"{val_rot_accuracy:.2f}")
            st.metric("Validation Inference Time (ms/image)", f"{val_avg_inference_time:.2f}")

        st.subheader("File Paths")
        st.json(file_paths)

    except Exception as e:
        st.error(f"An error occurred during training: {e}")
        st.exception(e) # Display full traceback for debugging
    finally:
        # Clear CUDA cache at the end of the run
        if torch.cuda.is_available():
            torch.cuda.empty_cache()
            st.info("CUDA cache cleared.")
